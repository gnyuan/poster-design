# -*- coding: utf-8 -*-
# @Time    : 2022/6/6 14:22
# @Author  : 臧成龙
# @FileName: usual.py
# @Software: PyCharm
# -*- coding: utf-8 -*-
import os
from datetime import datetime
from urllib.parse import unquote

import openpyxl
from django.http import FileResponse
from django.shortcuts import get_object_or_404
from django.db.models import Q
from fuadmin.settings import BASE_DIR, STATIC_URL
from ninja import Schema
from openpyxl import load_workbook

from .fu_auth import data_permission
from .fu_ninja import FuFilters
from .fu_response import FuResponse
from .usual import get_user_info_from_token


class ImportSchema(Schema):
    path: str


def create(request, data, model, using='default'):
    if not isinstance(data, dict):
        data = data.dict()
    # user_info = get_user_info_from_token(request)
    # # 创建时默认添加创建人、修改者和所属部门
    # data['creator_id'] = user_info['id']
    # data['modifier'] = user_info['name']
    # data['belong_dept'] = user_info['dept']
    query_set = model.objects.using(using).create(**data)
    return query_set


def batch_create(request, data, model, using='default'):
    user_info = get_user_info_from_token(request)
    data_list = []
    for item in data:
        if not isinstance(item, dict):
            item = item.dict()

        # 创建时默认添加创建人、修改者和所属部门
        item['creator_id'] = user_info['id']
        item['modifier'] = user_info['name']
        item['belong_dept'] = user_info['dept']
        data_list.append(model(**item))
    query_set = model.objects.using(using).bulk_create(data_list)
    return query_set


def delete(id, model, using='default'):
    instance = get_object_or_404(model.objects.using(using), id=id)
    instance.delete()
    pass


def update(request, id, data, model, using='default'):
    if not isinstance(data, dict):
        dict_data = data.dict()
    else:
        dict_data = data
    # user_info = get_user_info_from_token(request)
    # dict_data['modifier'] = user_info['name']
    instance = get_object_or_404(model.objects.using(using), id=id)
    for attr, value in dict_data.items():
        setattr(instance, attr, value)
    instance.save(using=using)
    return instance


def retrieve(request, model, filters: FuFilters = FuFilters(), using='default'):
    filters = data_permission(request, filters)
    if filters is not None:
        list_filters = Q()  # 构建一个执行列表过滤的过滤器
        field = None  # 如果有排序的话，会有这两个字段
        order = None
        # 将filters空字符串转换为None
        for attr, value in filters.__dict__.items():
            if attr == 'field':  # 如果字段支持排序，则自动加字段
                field = value
                setattr(filters, attr, None)
                continue
            if attr == 'order': # 如果字段支持排序，则自动加字段
                order = value
                setattr(filters, attr, None)
                continue
            if getattr(filters, attr) == '':
                setattr(filters, attr, None)
            if type(value) == list:
                if len(value) == 0:
                    setattr(filters, attr, None)
                else:
                    for item in value:
                        if item == '':
                            value.remove(item)
                    if len(value) == 0:
                        setattr(filters, attr, None)
                    else:
                        list_filters |= Q(**{attr: value})
                setattr(filters, attr, None)
        query_set = model.objects.using(using).filter(**filters.dict(exclude_none=True))
        query_set = query_set.filter(list_filters)
    else:
        query_set = model.objects.using(using).all()
    if order:  # 处理排序
        query_set = query_set.order_by(f'{field}' if order == 'ascend' else f'-{field}')
    return query_set


def export_data(request, model, scheme, export_fields):
    title_dict = {}
    for field in export_fields:
        field_obj = getattr(model, field).field
        title_dict[field_obj.column] = field_obj.help_text

    qs = retrieve(request, model)
    list_data = []
    for qs_item in qs:
        qs_item = scheme.from_orm(qs_item)
        dict_data = {}
        for item, value in title_dict.items():
            dict_data[value] = getattr(qs_item, item)
        list_data.append(dict_data)
    # 创建空的Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    for index, data in enumerate(list_data):
        if index == 0:
            ws.append(list(data.keys()))
        ws.append(list(data.values()))
    file_name = datetime.now().strftime('%Y%m%d%H%M%S%f') + '.xlsx'
    current_ymd = datetime.now().strftime('%Y%m%d')
    file_path = os.path.join(STATIC_URL, current_ymd)
    file_url = os.path.join(file_path, file_name)
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    wb.save(file_url)
    return FileResponse(open(file_url, "rb"), as_attachment=True)


def import_data(request, model, scheme, data, import_fields, using='default'):
    title_dict = {}
    for field in import_fields:
        field_obj = getattr(model, field).field
        title_dict[field_obj.help_text] = field_obj.column
    file_path = str(BASE_DIR) + unquote(data.path)
    wb = load_workbook(file_path)
    ws = wb.active
    title_value = []
    for index_row, row in enumerate(ws.values):
        if index_row == 0:
            title_value = row
        else:
            dict_data = {}
            for index, cell in enumerate(row):
                title_cell = title_value[index]
                value = title_dict.get(title_cell)
                if value is not None:
                    dict_data[value] = cell
            print(dict_data)
            data = scheme(**dict_data)
            create(request, data, model, using=using)
    return FuResponse(msg='导入成功')
